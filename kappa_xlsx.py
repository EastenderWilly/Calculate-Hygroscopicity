# include all functions from the math and numpy library
#import xlrd
#import xlwt
#from xlutils.copy import copy as xlutils_copy
from openpyxl import Workbook, load_workbook
from math import *
import math
import numpy as np
from scipy import fmin
import csv
import os
from os import listdir
from os.path import isfile, isdir, join
from scipy import optimize
import pandas as pd
from datetime import datetime
from datetime import timedelta
#import sysconfig
#print(sysconfig.__file__)
#print(sysconfig._init_posix.func_code)

DATA_DIR_CPC=input('Which folder are the CPC data at? (use correct case)\n')

if DATA_DIR_CPC=='':
   DATA_DIR_CPC = "./CPC" #default folder
   print('\nUse default folder: ',DATA_DIR_CPC)

DATA_DIR_CCN=input('\nWhich folder are the CCN data at? (please use correct case)\n')

if DATA_DIR_CCN=='':
   DATA_DIR_CCN = "./CCN" #default folder
   print('\nUse default folder: ',DATA_DIR_CCN)

DATA_DIR_SMPS=input('\nWhich folder are the SMPS data at? (please use correct case)\n')

if DATA_DIR_SMPS=='':
   DATA_DIR_SMPS = "./SMPS"#default folder
   print('\nUse default folder: ',DATA_DIR_SMPS)

SSr_calibr=['0.155905052','0.278318','0.49774','0.798991']

SSr_calibr_01=input('\nWhat is the calibrated value of SSr=0.1?    ')

try:
    SSr_calibr_01=float(SSr_calibr_01)
    SSr_calibr[0]=SSr_calibr_01
except:
    print('\nUse default value 0.155905052')

SSr_calibr_02=input('\nWhat is the calibrated value of SSr=0.2?    ')

try:
    SSr_calibr_02=float(SSr_calibr_02)
    SSr_calibr[1]=SSr_calibr_02
except:
    print('\nUse default value 0.278318')

SSr_calibr_03=input('\nWhat is the calibrated value of SSr=0.5?    ')

try:
    SSr_calibr_03=float(SSr_calibr_03)
    SSr_calibr[2]=SSr_calibr_03
except:
    print('\nUse default value 0.49774')

SSr_calibr_04=input('\nWhat is the calibrated value of SSr=0.8?    ')

try:
    SSr_calibr_04=float(SSr_calibr_04)
    SSr_calibr[3]=SSr_calibr_04
except:
    print('\nUse default value 0.798991')


#def iter_rows(ws):
#    for row in ws.iter_rows():
#        yield [cell.value for cell in row]

#All CPC data are in a single file
try:
 for filename in os.listdir(DATA_DIR_CPC):
  print ("Loading: %s" % filename)
  file_name=join(DATA_DIR_CPC,filename)
  date=[]
  CPC={'date_time':[],'CN':[]}
  try:
      cpc=load_workbook(filename=file_name)
      cpcreader=cpc.get_sheet_by_name('工作表1')
      line=1
      for row in cpcreader.iter_rows():
          if line>=2 and cpcreader['A%d'%(line)].value!='':
             #print(cpcreader['A%d'%(line)].value)
             CPC['CN'].append(cpcreader['B%d'%(line)].value)
             try:
                CPC['date_time'].append(cpcreader['A%d'%(line)].value)
                #CPC['date_time'].append(datetime.strptime(cpcreader['A%d'%(line)].value,"%d/%m/%Y %H:%M:%S"))
                #trandfer to time pattern
             except:
                CPC['date_time'].append('NaN')
          elif not row:
             continue
          line=line+1
  except IOError:
      print('\nFailed to open file ',file_name) 
 del date
except IOError:
 print('\nFailed to open folder ',DATA_DIR_CPC)

#print(CPC)


CCN={'date_time':[],'SSr':[],'stat':[],'CCN':[]}#data with stat=0 to be dropped 
try:
 for filename in os.listdir(DATA_DIR_CCN):
  print ("Loading: %s" % filename)
  file_name=join(DATA_DIR_CCN,filename)
  try:
      ccn=load_workbook(filename=file_name)
      ccnreader=ccn.get_sheet_by_name('工作表1')
      line=1
      for row in ccnreader.iter_rows():
          if line>=2 and ccnreader['A%d'%(line)].value!='': 
             try:
                CCN['date_time'].append(ccnreader['A%d'%(line)].value)
                #CCN['date_time'].append(datetime.strptime(ccnreader['A%d'%(line)].value,"%d/%m/%Y %H:%M:%S"))
             except:
                CCN['date_time'].append('NaN')

             if ccnreader['B%d'%(line)].value==0.1:
                CCN['SSr'].append(SSr_calibr[0])
             elif ccnreader['B%d'%(line)].value==0.2:
                CCN['SSr'].append(SSr_calibr[1])
             elif ccnreader['B%d'%(line)].value==0.5:
                CCN['SSr'].append(SSr_calibr[2])
             elif ccnreader['B%d'%(line)].value==0.8:
                CCN['SSr'].append(SSr_calibr[3])
             else:
                CCN['SSr'].append('NaN')

             CCN['stat'].append(ccnreader['C%d'%(line)].value)

             if ccnreader['C%d'%(line)].value == 1.0:                
                CCN['CCN'].append(ccnreader['D%d'%(line)].value)
             else:
                CCN['CCN'].append('NaN')
          elif not row:
             continue
          line=line+1
  except IOError:
      print('\nFailed to open file ',file_name) 
except IOError:
 print('\nFailed to open folder ',DATA_DIR_CCN)

#print(CCN)

diameter=[]
number=[]
SMPS={'date_time':[],'diameter':[],'number':[],'total number':[]}
CCN_CN_new={'date_time':[],'CCN':[],'CN':[],'SSr':[],'critical diameter':[],\
'kappa':[],'SMPS':[],'predict_kappa':[]}

try:
 for filename in os.listdir(DATA_DIR_SMPS):
  print ("Loading: %s" % filename)
  file_name=join(DATA_DIR_SMPS,filename)
  try:
      smps=load_workbook(filename=file_name)
      smpsreader=smps.get_sheet_by_name('工作表1')
      line=1
      for row in smpsreader.iter_rows():
          if line==1:
             for column in smpsreader.iter_rows(min_row=line,min_col=2,max_row=line,max_col=113):
                 for cell in column:
                     diameter.append(cell.value)
             #print(diameter)
          elif line>=2 and smpsreader['A%d'%(line)].value!='':
             try:
                date_time=smpsreader['A%d'%(line)].value
                if date_time.second!=0:
                   date_time=date_time-timedelta(seconds=date_time.second)
                date_time=datetime.strftime(date_time,"%d/%m/%Y %H:%M:%S")

             except:
                date_time='NaN'

             SMPS['date_time'].append(date_time)
             #print(date_time)
             SMPS['diameter'].append(diameter)
             for column in smpsreader.iter_rows(min_row=line,min_col=2,max_row=line,max_col=113):
                 for cell in column:
                     number.append(cell.value)

             SMPS['number'].append(number)
             SMPS['total number'].append(np.sum(number[1:112]))
            
             CCN_CN_new['date_time'].append(date_time)
             CCN_CN_new['SMPS'].append(np.sum(number[1:112]))


             #find where are the corresponding CCN and CPC data?

             date_time_CPC=[i for i,s in enumerate(CPC['date_time']) if date_time in s]
              
             date_time_CCN=[i for i,s in enumerate(CCN['date_time']) if date_time in s]
             
             if len(date_time_CCN)==1 and len(date_time_CPC)==1:
                for i in date_time_CCN:
                    for j in date_time_CPC:
                       CCN_CN_new['CCN'].append(CCN['CCN'][i])
                       CCN_CN_new['CN'].append(CPC['CN'][j])
                       try:
                          CCN_CN_ratio=float(CCN['CCN'][i])/float(CPC['CN'][j])
                       except:
                          CCN_CN_ratio='NaN'

                       CCN_CN_new['SSr'].append(CCN['SSr'][i])
                #print(CCN_CN_ratio) 
                if CCN_CN_ratio=='NaN':
                   continue
                elif CCN_CN_ratio>1.0 and CCN_CN_ratio<=1.1:
                   CCN_CN_ratio=1.0
                elif CCN_CN_ratio>=1.1 or CCN_CN_ratio<0.0:
                   CCN_CN_ratio='NaN'
               
                if np.sum(number[1:112])!='NaN' and CCN_CN_ratio!='NaN': #np.sum(row[1:112]):total number
                   critical_number=np.sum(number[1:112])*CCN_CN_ratio
                   for i in range(1,len(diameter)+1):
                       if i==len(diameter):
                          next_diameter=13.1
                       else:
                          next_diameter=float(diameter[-i-1])
                       critical_number=critical_number-float(number[112-i])
                       if critical_number==0.0:
                          lowerbound=(math.log10(float(diameter[-i]))+math.log10(next_diameter))/2.0
                          lowerbound=math.pow(10.0,lowerbound)
                          critical_diameter=lowerbound
                          break
                       elif critical_number<0.0:
                          upperbound=(math.log10(float(diameter[-i]))+math.log10(float(diameter[-i+1])))/2.0
                          upperbound=math.pow(10.0,upperbound)
                          lowerbound=(math.log10(float(diameter[-i]))+math.log10(next_diameter))/2.0
                          lowerbound=math.pow(10.0,lowerbound)
                          critical_diameter=(lowerbound*(float(number[112-i])+critical_number)\
+upperbound*(critical_number)*(-1.0))/float(number[112-i])
                          break
                       elif critical_number>0.0:
                          continue
                   CCN_CN_new['critical diameter'].append(critical_diameter) 
                else:
                   CCN_CN_new['critical diameter'].append('NaN')
                
             else:
                CCN_CN_new['CN'].append('NaN')
                CCN_CN_new['CCN'].append('NaN')
                CCN_CN_new['critical diameter'].append('NaN')
                CCN_CN_new['SSr'].append('NaN')
          elif not row:
             continue
          line=line+1
  except IOError:
      print('\nFailed to open file ',file_name)
except IOError:
 print('\nFailed to open folder ',DATA_DIR_SMPS)

#print(SMPS)


T=298.15
sigma=0.072
A=4.0*sigma*18.0/(8.314*T*1.0e6)
Dd=0.0


for i in range(0,len(CCN_CN_new['date_time'])):
    if CCN_CN_new['SSr'][i]!='NaN':
       d=float(CCN_CN_new['critical diameter'][i])*1.0e-9
       try:
          def g(k):
              def f(dwet):
                  s=-(dwet**3.0-d**3.0)/(dwet**3.0-d**3.0*(1.0-k))*exp(A/dwet)
                  return s
              dmin=optimize.fmin(f,d,disp=0)
              smax=(dmin**3.0-d**3.0)/(dmin**3.0-d**3.0*(1.0-k))*exp(A/dmin)
              t=abs(smax[0]-(1.0+float(CCN_CN_new['SSr'][i])*0.01))
              return t
          kmin=optimize.fmin(g,0.1,disp=0)
          kappa=kmin[0]
       
          CCN_CN_new['kappa'].append(kappa)
       except:
          CCN_CN_new['kappa'].append('NaN')
       kappa_temp=4.0*(A**3.0)/(27.0*(d**3.0)*(log(float(CCN_CN_new['SSr'][i])*0.01+1.0))**2.0)
       CCN_CN_new['predict_kappa'].append(kappa_temp)
    else:
       CCN_CN_new['kappa'].append('NaN')
       CCN_CN_new['predict_kappa'].append('NaN')


#write data into excel
with open('kappa.csv','w',newline='') as csvfile:
     data_writer=csv.writer(csvfile)
     header=['date_time','SSr='+SSr_calibr[0],'SSr='+SSr_calibr[1],\
'SSr'+SSr_calibr[2],'SSr'+SSr_calibr[3],'predict kappa']
     data_writer.writerow(header)
     for i in range(0,len(CCN_CN_new['CCN'])):
         if CCN_CN_new['SSr'][i]==SSr_calibr[0]:
            line=[CCN_CN_new['date_time'][i],CCN_CN_new['kappa'][i],'','','',CCN_CN_new['predict_kappa'][i]]
         elif CCN_CN_new['SSr'][i]==SSr_calibr[1]:
            line=[CCN_CN_new['date_time'][i],'',CCN_CN_new['kappa'][i],'','',CCN_CN_new['predict_kappa'][i]]
         elif CCN_CN_new['SSr'][i]==SSr_calibr[2]:
            line=[CCN_CN_new['date_time'][i],'','',CCN_CN_new['kappa'][i],'',CCN_CN_new['predict_kappa'][i]]
         elif CCN_CN_new['SSr'][i]==SSr_calibr[3]:
            line=[CCN_CN_new['date_time'][i],'','','',CCN_CN_new['kappa'][i],CCN_CN_new['predict_kappa'][i]]
         else:
            line=[CCN_CN_new['date_time'][i],'']
         data_writer.writerow(line)


with open('AR.csv','w',newline='') as csvfile:
     data_writer=csv.writer(csvfile)
     header=['date_time','SSr='+SSr_calibr[0],'SSr='+SSr_calibr[1],\
'SSr'+SSr_calibr[2],'SSr'+SSr_calibr[3]]
     data_writer.writerow(header)
     for i in range(0,len(CCN_CN_new['CCN'])):
         if CCN_CN_new['SSr'][i]==SSr_calibr[0]:
            line=[CCN_CN_new['date_time'][i],float(CCN_CN_new['CCN'][i])/float(CCN_CN_new['CN'][i]),'','','']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[1]:
            line=[CCN_CN_new['date_time'][i],'',float(CCN_CN_new['CCN'][i])/float(CCN_CN_new['CN'][i]),'','']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[2]:
            line=[CCN_CN_new['date_time'][i],'','',float(CCN_CN_new['CCN'][i])/float(CCN_CN_new['CN'][i]),'']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[3]:
            line=[CCN_CN_new['date_time'][i],'','','',float(CCN_CN_new['CCN'][i])/float(CCN_CN_new['CN'][i])]
         else:
            line=[CCN_CN_new['date_time'][i],'']
         data_writer.writerow(line)

with open('critical_diameter.csv','w',newline='') as csvfile:
     data_writer=csv.writer(csvfile)
     header=['date_time','SSr='+SSr_calibr[0],'SSr='+SSr_calibr[1],\
'SSr'+SSr_calibr[2],'SSr'+SSr_calibr[3]]
     data_writer.writerow(header)
     for i in range(0,len(CCN_CN_new['CCN'])):
         if CCN_CN_new['SSr'][i]==SSr_calibr[0]:
            line=[CCN_CN_new['date_time'][i],CCN_CN_new['critical diameter'][i],'','','']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[1]:
            line=[CCN_CN_new['date_time'][i],'',CCN_CN_new['critical diameter'][i],'','']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[2]:
            line=[CCN_CN_new['date_time'][i],'','',CCN_CN_new['critical diameter'][i],'']
         elif CCN_CN_new['SSr'][i]==SSr_calibr[3]:
            line=[CCN_CN_new['date_time'][i],'','','',CCN_CN_new['critical diameter'][i]]
         else:
            line=[CCN_CN_new['date_time'][i],'']
         data_writer.writerow(line)

